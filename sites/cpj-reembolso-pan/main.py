import os
import io
import json
import pdb
import sys
import base64
import re
import shutil
import zipfile
from datetime import datetime, time, timedelta
import traceback
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
import requests as _requests
import openpyxl


# Adiciona o caminho dos módulos
sys.path.append('C:\\www\\automacao')

# Importa funções do módulo CPJ API
from cpj_api import (
    set_api_credentials,
    api_logout,
    api_buscar_lancamentos_pan,
    api_buscar_documentos_spf,
    api_baixar_documento,
    formatar_data_lancamento
)

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

API_BASE_URL = 'https://app.leviatan.com.br/dcncadv/cpj/agnes'
API_LOGIN = 'api.teste5'
API_PASSWORD = 'dcnc2025'

NUMERO_CC_PAN = 1506

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_PATH = os.path.join(BASE_DIR, 'planilha')
JSON_PATH = os.path.join(PLANILHA_PATH, 'importados.json')
TAREFAS_PATH = os.path.join(BASE_DIR, 'tarefas.json')
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')
PDF_MERGE_PATH = os.path.join(BASE_DIR, 'pdf_merge')


# ============================================================================
# CARREGA PARÂMETROS (args ou config.json)
# ============================================================================

with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    _config = json.load(f)
NUMERO_RECIBO = _config.get('numero_recibo', '')
DATA_INICIAL_PESQUISA = datetime.strptime(_config['data_inicial'], "%Y-%m-%d") if _config.get('data_inicial') else None
DATA_FINAL_PESQUISA = datetime.strptime(_config['data_final'], "%Y-%m-%d") if _config.get('data_final') else None

# ============================================================================
# CONFIGURAÇÃO DA API
# ============================================================================

set_api_credentials(
    base_url=API_BASE_URL,
    login=API_LOGIN,
    password=API_PASSWORD,
    json_path=JSON_PATH,
    planilha_path=PLANILHA_PATH,
    pdf_merge_path=PDF_MERGE_PATH,
    config_json_path=CONFIG_PATH
)

# ============================================================================
# FUNÇÕES: TAREFAS.JSON
# ============================================================================

def carregar_tarefas():
    """Carrega o tarefas.json existente. Retorna None se não existir."""
    if not os.path.exists(TAREFAS_PATH):
        return None
    with open(TAREFAS_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def salvar_tarefas(dados: dict):
    """Salva o dicionário no tarefas.json."""
    with open(TAREFAS_PATH, 'w', encoding='utf-8') as f:
        json.dump(dados, f, indent=4, ensure_ascii=False)
    print(f'✓ tarefas.json salvo em: {TAREFAS_PATH}')


def limpar_pastas():
    """Apaga e recria as pastas de saída planilha e pdf_merge."""
    for pasta in [PLANILHA_PATH, PDF_MERGE_PATH]:
        if os.path.exists(pasta):
            shutil.rmtree(pasta)
        os.makedirs(pasta, exist_ok=True)
    print('✓ Pastas planilha e pdf_merge limpas.')


def carimbar_id_caso(pdf_path: str, id_caso: str, destino: str):
    """Adiciona o id_caso como badge no canto superior direito de todas as páginas do PDF."""
    try:
        reader = PdfReader(pdf_path)
        writer = PdfWriter()

        for page in reader.pages:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)

            packet = io.BytesIO()
            c = rl_canvas.Canvas(packet, pagesize=(w, h))

            font_size = 8
            c.setFont('Helvetica-Bold', font_size)
            texto = str(id_caso)
            text_width = c.stringWidth(texto, 'Helvetica-Bold', font_size)
            padding_x = 6
            padding_y = 4
            badge_w = text_width + padding_x * 2
            badge_h = font_size + padding_y * 2
            x = w - badge_w - 8
            y = h - badge_h - 8

            # Fundo escuro (badge)
            c.setFillColorRGB(0.15, 0.15, 0.15)
            c.roundRect(x, y, badge_w, badge_h, 3, fill=1, stroke=0)

            # Texto branco
            c.setFillColorRGB(1, 1, 1)
            c.drawString(x + padding_x, y + padding_y, texto)
            c.save()
            packet.seek(0)

            overlay = PdfReader(packet).pages[0]
            page.merge_page(overlay)
            writer.add_page(page)

        with open(destino, 'wb') as f_out:
            writer.write(f_out)
    except Exception as e:
        print(f'  ⚠ Erro ao carimbar PDF: {e}. Usando original.')
        shutil.copy2(pdf_path, destino)


def obter_dados_lancamentos_cpj():
    """
    Busca lançamentos PAN via API e salva em tarefas.json.

    Fluxo:
        1. Login na API CPJ Agnes
        2. api_buscar_lancamentos com numero_cc=1506

    Returns:
        dict: Estrutura completa
    """

    print('\n' + '='*60)
    print('OBTENDO DADOS DE LANÇAMENTOS - BANCO PAN')
    print('='*60)

    # 1. Busca lançamentos na API
    print('\n[1/2] Buscando lançamentos na API...')
    lancamentos = api_buscar_lancamentos_pan(
        data_inicial=DATA_INICIAL_PESQUISA,
        data_final=DATA_FINAL_PESQUISA,
        numero_cc=NUMERO_CC_PAN,
        limit=1000,
        titulo=NUMERO_RECIBO,
        valor_limite_original=False
    )

    return lancamentos


def baixar_e_mesclar_documentos(lancamentos: list):
    """
    Baixa todos os documentos de todas as tarefas e gera um único PDF mesclado.

    Args:
        lancamentos: lista de dicts com ao menos 'id_spf'
    """
    print('\n' + '='*60)
    print('BAIXANDO E MESCLANDO DOCUMENTOS')
    print('='*60)

    os.makedirs(PDF_MERGE_PATH, exist_ok=True)

    pdf_temp = os.path.join(BASE_DIR, 'pdf_temporarios')
    os.makedirs(pdf_temp, exist_ok=True)

    merger = PdfMerger()
    total_adicionados = 0
    lancamentos_totais = []

    for idx, lancamento in enumerate(lancamentos, start=1):
        
        lancamento['tipo_pagamento'] = 'Boleto'
        lancamento['numero_guia'] = 'NAO_ENCONTRADA'
        lancamento_unico = lancamento
        identificou_comprovante = False
        identificou_guia = False

        id_spf = lancamento.get('id_spf', '')
        id_caso = lancamento.get('id_caso')
        numero_integracao = lancamento.get('id_caso', '')
        #pdb.set_trace()  # Debug: Verificar dados do lançamento antes de processar os documentos

        print(f'\n[{idx}/{len(lancamentos)}] SPF {id_spf} | {id_caso}')

        if not id_spf:
            print('  ⚠ id_spf vazio, pulando...')
            pdb.set_trace()  # Debug: Verificar lançamento com id_spf vazio
            continue
        
        # if id_spf != '51506':
        #     continue
        
        # 1. Busca documentos do SPF
        documentos = api_buscar_documentos_spf(id_spf)
        
        if not documentos:
            print(f'  ⚠ Nenhum documento encontrado para SPF {id_spf}')
            pdb.set_trace()  # Debug: Verificar lançamento que não retornou documentos
            continue

        print(f'  ✓ {len(documentos)} documento(s) encontrado(s)')

        if len(documentos) == 2:
            documentos = sorted(documentos, key=lambda x: x['data_criacao'], reverse=False)

        documentos_validos = []

        # 2. Baixa e adiciona cada documento ao merger
        for indicedoc, doc in enumerate(documentos, start=1):

            # if identificou_comprovante and identificou_guia:
            #     print('  ✓ Comprovante e guia já identificados, pulando documentos restantes...')
            #     break
            
            id_ged = doc.get('id_ged')
            # ocr_realizada = False

            if not id_ged:
                continue

            destino = os.path.join(pdf_temp, f'{id_caso}_{id_spf}_{id_ged}_{idx}_{indicedoc}.pdf')
            sucesso = api_baixar_documento(id_ged, destino)

            # if '564055' in numero_integracao:
            #     pdb.set_trace()  # Debug: Verificar lançamento específico com número de integração 564055

            if sucesso and os.path.exists(destino):
                
                try:
                    reader = PdfReader(destino)
                    texto_pdf = '\n'.join(
                        page.extract_text() or '' for page in reader.pages
                    )                        

                    with open(destino, 'rb') as f_pdf:
                        b64 = base64.b64encode(f_pdf.read()).decode('utf-8')

                    if texto_pdf.strip() == "":
                        print(f'  ⚠ Documento {id_ged} sem texto extraído. Executando OCR N8N...')
                        try:
                            resp_ocr = _requests.post(
                                'https://n8n-diascosta.up.railway.app/webhook/b1292c39-06a5-49e6-8439-c816a899767a',
                                json={'base64': b64},
                                timeout=500
                            )

                            if resp_ocr.status_code == 200:
                                texto_pdf = json.loads(resp_ocr.text).get('texto_ocr', '')
                                print(f'  ✓ OCR retornou {len(texto_pdf)} caracteres para documento {id_ged}.')
                                ocr_realizada = True
                            else:
                                print(f'  ⚠ OCR retornou status {resp_ocr.status_code} para documento {id_ged}.')
                                pdb.set_trace()  # Debug: Verificar erro do OCR 1
                        except Exception as e_ocr:
                            pdb.set_trace()  # Debug: Verificar erro do OCR 2
                            print(f'  ✗ Erro ao executar OCR no documento {id_ged}: {e_ocr}')

                    total_adicionados += 1
                    
                    if ((('btg' in texto_pdf.lower() or 'útil' in texto_pdf.lower()) and 'comprovante' in texto_pdf.lower())
                        or ('comprovante' in doc.get('path').lower()) 
                        or 'comprovante de pagamento' in texto_pdf.lower() 
                        or 'comprovante de pagamento de titulos' in texto_pdf.lower() 
                        or 'comprovante pix' in texto_pdf.lower()) and identificou_comprovante == False and 'guia' not in doc.get('path').lower():
                        
                        merger.append(destino)
                        identificou_comprovante = True
                        if 'pix' in texto_pdf.lower():
                            print(f'  ✓ Documento {id_ged} contém PIX.')
                            lancamento['tipo_pagamento'] = 'PIX'
                        elif 'boleto' in texto_pdf.lower() or 'linha digitável' in texto_pdf.lower():
                            print(f'  ✓ Documento {id_ged} contém Boleto.')
                            lancamento['tipo_pagamento'] = 'Boleto'
                        else:

                            if 'útil' in texto_pdf.lower():
                                print(f'  ✓ Documento {id_ged} contém "útil", indicando possível comprovante de pagamento do BTG.')
                                lancamento['tipo_pagamento'] = 'Boleto'

                            elif 'comprovante de pagamento de titulos' in texto_pdf.lower():
                                print(f'  ✓ Documento {id_ged} contém "COMPROVANTE DE PAGAMENTO DE TITULOS", indicando possível comprovante de pagamento do BTG.')
                                lancamento['tipo_pagamento'] = 'Boleto'

                            elif 'comprovante de pagamento' in texto_pdf.lower():
                                print(f'  ✓ Documento {id_ged} contém "COMPROVANTE DE PAGAMENTO", indicando possível comprovante de pagamento do BTG.')
                                lancamento['tipo_pagamento'] = 'Boleto'
                            
                            else:
                                pdb.set_trace()  # Debug: Verificar documento que contém "comprovante" mas não identificou tipo de pagamento
                        
                        #lancamento['path_documento'] = doc['path']
                        documentos_validos.append({'tipo_pagamento': lancamento['tipo_pagamento']})
                        continue

                    else:

                        destino_carimbado = destino.replace('.pdf', '_c.pdf')
                        carimbar_id_caso(destino, str(id_caso), destino_carimbado)
                        merger.append(destino_carimbado)

                        # Envia texto para validar e obter número da guia
                        try:
                            for _tentativa in range(5):
                                resp_guia = _requests.post(
                                    'https://n8n-diascosta.up.railway.app/webhook/71b2da26-7510-4415-bdb8-c156ad5f03e2-valida-retorna-guia',
                                    json={'texto': texto_pdf.lower(), 'base64': b64, 'numero_processo': lancamento.get('numero_processo', '')},
                                    timeout=130
                                )
                                if resp_guia.status_code != 500:
                                    break
                                print(f'  ⚠ Guia retornou 500, tentativa {_tentativa + 1}/5. Tentando novamente...')
                                time.sleep(5)

                            # Debug: Verificar resposta da validação da guia
                            if resp_guia.status_code == 200:
                                dados_guia = resp_guia.json()
                                if dados_guia.get('guia_encontrada'):
                                    identificou_guia = True
                                    lancamento['numero_guia'] = dados_guia['numero_guia']
                                    print(f'  ✓ Guia encontrada: {lancamento["numero_guia"]}')
                                else:
                                    pdb.set_trace()  # Debug: Verificar resposta da API quando guia não é encontrada
                                    print(f'  ⚠ Guia não encontrada no documento {id_ged}')
                            else:
                                print(f'  ⚠ Erro na validação da guia. Status: {resp_guia.status_code}')
                                lancamento['numero_guia'] = "NUMERO_GUIA_NAO_ENCONTRADA_1"

                        except Exception as e_guia:
                            print(f'  ✗ Erro ao validar guia: {e_guia}')
                            lancamento['numero_guia'] = "NUMERO_GUIA_NAO_ENCONTRADA_2"

                        #lancamento['path_documento'] = doc['path']

                        # if len(documentos) == 2:
                        #     lancamentos_totais.append(lancamento.copy())
                    
                        documentos_validos.append({'numero_guia': lancamento['numero_guia']})

                    # if len(documentos) >= 2:
                    #     lancamentos_totais.append(lancamento.copy())

                except Exception as e:
                    print(f'  ✗ Erro ao adicionar PDF {id_ged}: {e}')
                    pdb.set_trace()  # Debug: Verificar erro ao processar documento específico

        #caso tenha mais de 2 documentos, adiciona o lançamento para cada guia encontrada, concatenando os números das guias e tipos de pagamento
        if len(documentos) > 2:
            pdb.set_trace()  # Debug: Verificar lançamento com mais de 2 documentos e como estão sendo concatenados os dados
            lancamento.pop('numero_guia', None)
            lancamento.pop('tipo_pagamento', None)

            #pdb.set_trace()  # Debug: Verificar lançamento com mais de 2 documentos e como estão sendo concatenados os dados
            lancamento['valor_original'] = lancamento['valores_parcelas']

            for i in range(0, int(len(documentos_validos)/2)):
                for key in ['numero_guia', 'tipo_pagamento']:
                    valores = [d[key] for d in documentos_validos if key in d]
                    lancamento[f'{key}'] = ','.join(valores)
                lancamentos_totais.append(lancamento.copy())

        if not identificou_comprovante:
            print(f'  ⚠ Atenção: Para SPF {id_spf}, não identificou comprovante. Documentos encontrados: {len(documentos)}')
            #pdb.set_trace()  # Debug: Verificar detalhes do lançamento que não identificou comprovante

        if not identificou_guia:
            print(f'  ⚠ Atenção: Para SPF {id_spf}, não identificou guia. Documentos encontrados: {len(documentos)}')
            #pdb.set_trace()  # Debug: Verificar detalhes do lançamento que não identificou guia

        lancamentos_totais.append(lancamento.copy())

    # 3. Salva o único PDF final
    if total_adicionados > 0:
        nome_final = f'pan_{DATA_INICIAL_PESQUISA.strftime("%Y%m%d")}_{DATA_FINAL_PESQUISA.strftime("%Y%m%d")}.pdf'
        destino_final = os.path.join(PDF_MERGE_PATH, nome_final)
        try:
            merger.write(destino_final)
            print(f'\n✓ PDF final salvo: {destino_final}')
            print(f'✓ Total de arquivos mesclados: {total_adicionados}')
        except Exception as e:
            print(f'\n✗ Erro ao salvar PDF final: {e}')
            traceback.print_exc()
    else:
        print('\n⚠ Nenhum PDF foi baixado.')

    merger.close()

    # Limpa pasta temporária
    shutil.rmtree(pdf_temp, ignore_errors=True)
    print('✓ Download e merge concluídos.')

    return lancamentos_totais


def preencher_planilha(lancamentos: list):
    """
    Copia a planilha modelo e preenche com os dados dos lançamentos.

    Args:
        lancamentos: lista de dicts com dados dos lançamentos
    """
    print('\n' + '='*60)
    print('PREENCHENDO PLANILHA')
    print('='*60)

    modelo_path = os.path.join(BASE_DIR, 'planilha_modelo', 'pan', 'planilha_padrao.xlsx')
    os.makedirs(PLANILHA_PATH, exist_ok=True)
    nome_saida = f'pan_{DATA_INICIAL_PESQUISA.strftime("%Y%m%d")}_{DATA_FINAL_PESQUISA.strftime("%Y%m%d")}.xlsx'
    destino_path = os.path.join(PLANILHA_PATH, nome_saida)

    shutil.copy2(modelo_path, destino_path)
    print(f'✓ Planilha modelo copiada para: {destino_path}')

    wb = openpyxl.load_workbook(destino_path, keep_links=False)
    ws = wb.active

    def formatar_valor(valor):
        try:
            f = float(valor)
            total_centavos = round(f * 100)
            inteiro = total_centavos // 100
            centavos = total_centavos % 100
            inteiro_fmt = f'{inteiro:,}'.replace(',', '.')
            return f'{inteiro_fmt},{centavos:02d}'
        except (TypeError, ValueError):
            return str(valor) if valor is not None else ''
    #pdb.set_trace()  # Debug: Verificar dados dos lançamentos antes de preencher a planilha
    for idx, lancamento in enumerate(lancamentos):
        linha = idx + 3  # linhas 1 e 2 são cabeçalhos

        ws[f'B{linha}'] = lancamento.get('id_caso', '')
        ws[f'C{linha}'] = formatar_data_lancamento(lancamento.get('data_lancamento', ''))
        ws[f'D{linha}'] = formatar_valor(lancamento.get('valor_original', ''))
        ws[f'E{linha}'] = lancamento.get('numero_guia', '')
        ws[f'F{linha}'] = lancamento.get('tipo_despesa', '')
        ws[f'H{linha}'] = 'Guia (Boleto)'
        ws[f'I{linha}'] = lancamento.get('tipo_pagamento', '')
        ws[f'M{linha}'] = lancamento.get('numero_processo', '')
        ws[f'N{linha}'] = lancamento.get('autor_nome', '')
        ws[f'O{linha}'] = 'BANCO PAN'
        ws[f'P{linha}'] = 'BANCO PAN'

        print(f'  ✓ Linha {linha}: SPF {lancamento.get("id_spf", "")} | {lancamento.get("autor_nome", "")}')

    wb.save(destino_path)

    # -------------------------------------------------------------------------
    # Correção de tabelas via zipfile:
    # openpyxl corrompe o XML das tabelas ao serializar. Restauramos os
    # arquivos de tabela com os XMLs originais do modelo (sem modificação).
    # -------------------------------------------------------------------------
    with zipfile.ZipFile(modelo_path, 'r') as zm:
        tabelas_modelo = {
            n: zm.read(n) for n in zm.namelist()
            if re.match(r'xl/tables/table\d+\.xml', n)
        }

    with zipfile.ZipFile(destino_path, 'r') as zd:
        conteudos = {n: zd.read(n) for n in zd.namelist()}

    for nome, conteudo in tabelas_modelo.items():
        conteudos[nome] = conteudo
        print(f'  ✓ Tabela {nome} restaurada do modelo')

    with zipfile.ZipFile(destino_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for nome, dados in conteudos.items():
            zout.writestr(nome, dados)

    print(f'\n✓ Planilha salva: {destino_path}')
    print(f'✓ Total de linhas preenchidas: {len(lancamentos)}')



def zerar_config():
    """Zera os valores do arquivo config.json após submissão bem-sucedida"""
    try:
        print('\nZerando config.json...')
        
        config_vazio = {
            "numero_recibo": "",
            "data_inicial": "",
            "data_final": "",
            "iniciado_em": ""
        }
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config_vazio, f, indent=4, ensure_ascii=False)
        print('✓ Config.json zerado com sucesso!')
        return True
    except Exception as e:
        print(f'✗ Erro ao zerar config.json: {e}')
        return False
# ============================================================================
# MAIN
# ============================================================================

def main():

    limpar_pastas()

    integracao = True

    if integracao:
        lancamentos = obter_dados_lancamentos_cpj()

        if not lancamentos:
            print('✗ Nenhum lançamento encontrado. Encerrando.')
            return

        salvar_tarefas(lancamentos)
    
    else:
        lancamentos = carregar_tarefas()
        if not lancamentos:
            print('✗ Nenhum dado encontrado em tarefas.json. Execute a integração primeiro.')
            return
        print(f'✓ {len(lancamentos)} lançamentos carregados de tarefas.json')

    lancamentos = baixar_e_mesclar_documentos(lancamentos)

    preencher_planilha(lancamentos)

    api_logout()

    zerar_config()

if __name__ == '__main__':
    main()
